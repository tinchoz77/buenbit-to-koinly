from Koinly import *
import DolarMEP as mep
import openpyxl

report = KoinlyCustomFile()
mep = mep.DolarMEP()

# Lists expected operation types and if the operation is reported in two lines
BUENBIT_OPERATIONS = {"CONVERSION": True,
                      "DEPOSITO": False,
                      "RETIRO": False,
                      "TRANSFERENCIA P2P RECIBIDA": False,
                      "CONSUMO DE TARJETA": False,
                      "AJUSTE DE TARJETA": False,
                      "CASHBACK": False}

# Define variable to load the buenbit report
xls = openpyxl.load_workbook("activity_report.xlsx").active

# Loop through every transaction (assuming there's not any failed trx)
# TODO: Check for FAILED status 
def check_ars_mep(trx):
    if trx.sent_currency == "ID:3622":
        valor_mep = mep.find_mep(trx.date)
        if valor_mep is not None:
            trx.net_worth_amount = trx.sent_amount / valor_mep
            trx.net_worth_currency = "USD"
    if trx.received_currency == "ID:3622":
        valor_mep = mep.find_mep(trx.date)
        if valor_mep is not None:
            trx.net_worth_amount = trx.received_amount / valor_mep
            trx.net_worth_currency = "USD"

for row_num in range(2, xls.max_row+1):
    print(f"\nLine {row_num}: ", end="")
    add_trx = False
    xls_moneda = xls.cell(row=row_num, column=5).value
    xls_monto = xls.cell(row=row_num, column=6).value
    xls_costored = xls.cell(row=row_num, column=7).value
    xls_txid = xls.cell(row=row_num, column=8).value

    if xls.cell(row=row_num, column=1).value is None:
        # complete previous transaction
        trx.set_amount(xls_monto, xls_moneda)
        trx.set_cost(xls_costored, xls_moneda)

        print(f"+--> completing Id: {xls_id} - Operacion: {last_operation}", end="")
        add_trx = True
    else:
        # new transaction
        xls_date = xls.cell(row=row_num, column=1).value
        xls_id = xls.cell(row=row_num, column=2).value
        xls_operation = xls.cell(row=row_num, column=3).value
        last_operation = xls_operation

        if not xls_operation in BUENBIT_OPERATIONS:
            raise Exception(f"Operation {xls_operation} not defined")
        
        print(f"Id: {xls_id} - Operacion: {xls_operation}", end="")
        trx = KoinlyCustomTransaction(xls_id, xls_date, xls_operation, xls_txid)
        trx.set_amount(xls_monto, xls_moneda)
        trx.set_cost(xls_costored, xls_moneda)

        add_trx = not BUENBIT_OPERATIONS[xls_operation]

    if add_trx:
        print(" "+u'\u2713', end="")
        check_ars_mep(trx)
        report.add_trx(trx)

# Export koinly csv file
print("\n\nWriting file...", end="")
report.write("buenbit_koinly_report.csv")
print(" done.")